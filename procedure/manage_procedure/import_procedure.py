#!/usr/bin/python
# -*- coding: UTF-8 -*-
"""
    import procedure_file
"""
import xlrd,json,re
from utils import core
from collections import Counter
import shutil
import os
import xlrd
from utils.ClientModels import Procedure, Usecase, Phrase
import openpyxl
import re

Formula_Regular = r'^([\w]+)=([\+\-]?[\d]+([\.][\d]*)?([Ee][+-]?[\d]+)?)'
zhmodel = re.compile(u'[\u4e00-\u9fa5]')


def select_p():
    list_p = list_procedure()
    if os.path.exists(list_p[0]):
        parse_procedure(list_p[0])
    else:
        print("解析的规程不存在")


def open_excel(pstr):
    try:
        procedure = xlrd.open_workbook(pstr)
        return procedure
    except:
        print(str(e))

def parse_procedure(pstr, wordPath):
    global all_formula_val  # 全局变量用来保存规程的变量名
    all_formula_val = []
    data = open_excel(pstr)
    excel = data.sheets()[0]
    procedure = Procedure()
    procedure.name = excel.row_values(0)[1]
    procedure.number = excel.row_values(0)[3]
    procedure.usecase = []
    procedure.path = pstr
    procedure.wordPath = wordPath
    usecase_save = parse_usecase(excel, procedure)
    procedure.usecase = json.dumps(procedure.usecase)
    procedure.type = excel.row_values(0)[5]
    if not procedure.name:
        procedure.name = os.path.basename(pstr)
    procedure.save_obj()
    return procedure.name


def parse_usecase(excel, procedure):
    usercase_pass = True
    for rownum in range(1, excel.nrows):
        row = excel.row_values(rownum)
        if row[0] == "测试用例":
            usecase = Usecase()
            usecase.name = row[1]
            usecase.number = row[3]
            procedure.usecase.append(usecase.number)
            usecase.IC = row[8]
            usecase.description = row[5]
            usecase.operation = json.dumps(parse_operation(excel, rownum))
            # print(json.loads(usecase.operation))
            if not usecase.name:
                usecase.name = usecase.number
            usecase.save_obj()
    return usercase_pass


def parse_operation(excel, rownum):
    oper = []
    k = 0
    subrownum = rownum + 3
    total_lines = excel.nrows
    while subrownum < excel.nrows:
        # print(subrownum)
        subrow = excel.row_values(subrownum)
        # print(subrow)
        # print(type(subrow[1]), subrow[1])
        # print(subrow[1])
        if subrow[0] == "测试用例":
            break
        try:
            # int(subrow[0])
            float(subrow[0])
        except Exception as e:
            # print(e)
            # print(subrow, 11111111111111111111111)
            oper.append([[subrow[0],subrow[1]]])
            j = 1
            k += 1
            subrownum += 1
            # print(1, subrownum)
            continue

        if "等待" in subrow[1]:
            print(subrow[1])
            write_formula = parse_write(subrow[1])
            oper[k - 1].append(["wait", {"sort": str(int(subrow[0])), "name": subrow[1], "remark": subrow[8],
                                          # "startdelay": parse_startdelay(subrow[1]),
                                          # "delay": parse_delay(subrow[1]),
                                          "location": subrow[2], "formula": write_formula}])
            j = j + 1
            subrownum += 1
        elif "设置" in subrow[1] and re.compile('[a-zA-Z]').search(subrow[1]):
            # print('okkk')
            write_formula = parse_write(subrow[1])
            oper[k - 1].append(["WRITE", {"sort": str(int(subrow[0])), "name": subrow[1], "remark": subrow[8],
                                          # "startdelay": parse_startdelay(subrow[1]),
                                          # "delay": parse_delay(subrow[1]),
                                          "location": subrow[2], "formula": write_formula}])
            j = j + 1
            subrownum += 1
            # print(2, subrownum)
        elif "收到" in subrow[1] and re.compile('[a-zA-Z]').search(subrow[1]):
            # print('okkk')
            write_formula = parse_write(subrow[1])
            oper[k - 1].append(["wait", {"sort": str(int(subrow[0])), "name": subrow[1], "remark": subrow[8],
                                          # "startdelay": parse_startdelay(subrow[1]),
                                          # "delay": re.findall('\d+',subrow[1])[0],
                                          "location": subrow[2], "formula": write_formula}])
            j = j + 1
            subrownum += 1
        elif "接收" in subrow[1]:
            # print('okkk')
            write_formula = parse_write(subrow[1])
            oper[k - 1].append(["wait", {"sort": str(int(subrow[0])), "name": subrow[1], "remark": subrow[8],
                                          # "startdelay": parse_startdelay(subrow[1]),
                                          # "delay": re.findall('\d+',subrow[1])[0],
                                          "location": subrow[2], "formula": write_formula}])
            j = j + 1
            subrownum += 1
            # print(3, subrownum)
        elif "检查" in subrow[1] and re.compile('[a-zA-Z]').search(subrow[1]):
            first_read = True
            while subrownum < excel.nrows:
                rsubrow = excel.row_values(subrownum)
                if '检查' not in rsubrow[1]:
                    break
                if first_read:
                    read_formula = parse_read(rsubrow[3])
                    oper[k - 1].append(
                        ["READ", {"sort": str(int(rsubrow[0])), "name": rsubrow[1], "remark": rsubrow[8],
                                  # "delay": parse_delay(rsubrow[1]), "except": rsubrow[3],
                                  "location": rsubrow[2], "formula": read_formula}
                         ])
                elif not first_read:
                    read_formula = parse_read(rsubrow[3])
                    oper[k - 1][j].append({"sort": str(int(rsubrow[0])), "name": rsubrow[1], "remark": rsubrow[8],
                                           # "delay": parse_delay(rsubrow[1]), "except": rsubrow[3],
                                           "location": rsubrow[2], "formula": read_formula})
                subrownum += 1
            j = j + 1
        # elif "初始化" in subrow[1]:
        #     oper[k - 1].append(["INIT", {"sort": str(int(subrow[0])), "name": subrow[1], "remark": subrow[8],
        #                                  "location": subrow[2], "except": subrow[3]}
        #                         ])
        #     j = j + 1
        #     subrownum += 1
        # elif not zhmodel.search(subrow[1]):
        #     # print(subrow[1],'111')
        #     oper[k - 1].append(["WRITE", {"sort": str(int(subrow[0])), "name": subrow[1], "remark": subrow[8],
        #                                   # "startdelay": parse_startdelay(subrow[1]),
        #                                   # "delay": parse_delay(subrow[1]),
        #                                   "location": subrow[2], "formula": write_formula}])
        #     j = j + 1
        #     subrownum += 1
        else:
            # print(subrow)
            # # continue
            # # oper[k - 1].append(["INIT", {"sort": str(int(subrow[0])), "name": subrow[1], "remark": subrow[8],
            # #                              "location": subrow[2], "except": subrow[3]}])
            # write_formula = parse_write(subrow[1])
            # oper[k - 1].append(["WRITE", {"sort": str(int(subrow[0])), "name": subrow[1], "remark": subrow[8],
            #                               # "startdelay": parse_startdelay(subrow[1]),
            #                               # "delay": parse_delay(subrow[1]),
            #                               "location": subrow[2], "formula": write_formula}])
            # j = j + 1
            subrownum += 1
            print(4, subrownum)
    return oper


# 解析设置值的公式，匹配顺序是科学计数法-》小数-》整数
def parse_write(name):
    formula = []
    for subline in name.replace(" ", "").split("\n"):
        if subline != "" and "DELAY" not in subline:
            match_result = re.match(Formula_Regular, subline)
            if match_result:
                formula.append(match_result.group())
                # 保存全部的变量名 供后续检查变量是否存在
                all_formula_val.append(match_result.group().split("=")[0])
    # 如果设置公式匹配数为0 ，返回fasle
    if len(formula) == 0:
        return False
    return ",".join(formula)


# 解析检查值的公式,匹配顺序是科学计数发 -》小数-》整数
def parse_read(name):
    assert_line = name.split("\n")
    # print(assert_line)
    if len(assert_line) > 0 and '' not in assert_line:
        formula = []
        formula.append(re.match('^\\w+', assert_line[0]).group())
        for i in range(1, len(assert_line)):
            match_result = re.match(Formula_Regular, assert_line[i])
            if match_result:
                formula.append(match_result.group())
                # 保存全部的变量名 供后续检查变量是否存在
                all_formula_val.append(match_result.group().split("=")[0])
    else:
        return False
    if len(formula) < 2:  # 预期结果的公式匹配错误
        return False
    else:
        return ",".join(formula)


# 解析公式中的delaytime,比如DELAY4=4s 则返回 "DELAY4,4"
def parse_delay(name):
    line = name.split("\n")
    delay_time = ""
    for subline in line:
        if "DELAY" in subline:
            if re.match(r'^(DELAY\d+)', subline):
                delay_time = ["START" + subline.split("=")[0], re.match(r'\d+', subline.split("=")[1]).group()]
    return ",".join(delay_time)


# 解析公式中的startdelay,比如是STARTDELAY4=0 ,则返回STARTDELAY4
def parse_startdelay(name):
    line = name.split("\n")
    startdelay = ""
    for subline in line:
        if "STARTDELAY" in subline:
            startdelay = subline.split("=")[0].replace(" ", "")
    return startdelay

def check_excel_line(excel):
    excel_error = []
    if (excel.row_values(0)[0] != '规程名称') or (excel.row_values(0)[2] != '规程编号'):
        excel_error.append("不是规程文件")
        return excel_error
    for i in range(0, excel.nrows):
        if (not excel.row_values(i)[0]) and (not excel.row_values(i)[1]):
            excel_error.append('表格第' + str(i + 1) + "行,规程可能有空行")
        if type(excel.row_values(i)[0]) is int:
            if "检查" in excel.row_values(i)[1]:
                if not parse_read(excel.row_values(i)[3]):
                    excel_error.append('表格第' + str(i + 1) + "行,公式错误")
            elif "设置" in excel.row_values(i)[1]:
                if not parse_write(excel.row_values(i)[1]):
                    excel_error.append('表格第' + str(i + 1) + "行,公式错误")
            elif "初始化" in excel.row_values(i)[1]:
                pass
            else:
                excel_error.append('表格第' + str(i + 1) + "行,无法解析该语义")
    return excel_error


class DCSLogType(object):
    DEFAULT = 'default'
    INFO = 'info'
    WARNING = 'warning'
    ERROR = 'error'
    RIGHT = 'right'

    @classmethod
    def types(cls):
        return [
            cls.DEFAULT, cls.INFO, cls.WARNING, cls.ERROR, cls.RIGHT
        ]

def judgeCN(con):
    # 判断字符串是否包含中文
    if re.compile(u'[\u4e00-\u9fa5]').search(con):
        return True
    else:
        return False

def getOprType(action):
    for key, value in phraseDic.items():
        if key in action:
            if value == 'SET' or not judgeCN(action) and '=' in action:
                return value
            elif value in 'CHECK' and re.sub('[\u4e00-\u9fa5]', '', action):
                return value
            else:
                return value



def parse_GS_procedure(excelpath):
    titleDic = {}
    pharses = Phrase.get_all()
    for phrase in pharses:
      phraseDic[phrase.name] = phrase.operation 
    wb = openpyxl.load_workbook(excelpath)
    ws = wb.active
    Program = {}
    info = {}
    order = []
    procedure = Procedure()
    procedure.name = ws[1][1].value
    procedure.number = ws[1][3].value
    procedure.usecase = []
    procedure.path = excelpath
    Program['sections'] = []
    info['procedureName'] = ws[1][1].value
    info['procedureNumber'] = ws[1][3].value
    info['usecaseName'] = ws[2][1].value
    info['usecaseNumber'] = ws[1][3].value
    info['IC'] = ws[2][8].value
    Program['info'] = info
    for index, x in enumerate(list(ws.iter_rows())[2]):
        if x.value:
            titleDic[x.value] = index
    titleDic['操作位置'] = titleDic['试验步骤'] + 1
    for i, row in enumerate(list(ws.iter_rows())[4:]):
        if row[0].value == "测试用例":
            usecase = Usecase()
            usecase.name = row[1].value
            usecase.number = row[3].value
            procedure.usecase.append(usecase.number)
            usecase.IC = row[8].value
            usecase.description = row[5].value
        # try:
        action = str(row[titleDic['试验步骤']].value)
        index = row[titleDic['序号']].value
        if 'STEP' in str(index) and judgeCN(str(action)):
            try:
                sectionDic['oprs'] = oprs
                Program['sections'].append(sectionDic)
            except:
                pass
            sectionDic = {}
            oprs = []
            sectionDic['STEP'] = re.findall('\d+',index)[0]
            order.append(sectionDic['STEP'])
            sectionDic['name'] = action
            # print(sectionDic)
            continue
        oprDict = {}
        oprDict['type'] = getOprType(action)
        if oprDict['type'] == 'SET':
            oprDict['opr'] = []
            for write in action.split('\n')[1:]:
                oprDict['opr'].append(write.split('='))
            # 存储二维数组 [[变量名,设定值]...]
        elif oprDict['type'] == 'CHECK':
            oprDict['opr'] = re.sub('[\u4e00-\u9fa5]', '', action)
            oprDict['expectResult'] = str(row[titleDic['预期结果']].value)
        # elif oprDict['type'] == 'CALL':
        #   oprDict['opr'] = 
        elif not oprDict['type']:
            continue
        oprDict['lineNumber'] = i
        oprDict['position'] = row[titleDic['操作位置']].value
        oprDict['remark'] = row[titleDic['备注']].value
        oprs.append(oprDict)
        # except:
        #   print(filePath, [x.value for x in row])
    sectionDic['oprs'] = oprs
    Program['sections'].append(sectionDic)
    # sectionDic = {}
    # oprs = []
    # sectionDic['STEP'] = re.findall('\d+',index)[0]
    # order.append(sectionDic['STEP'])
    # sectionDic['name'] = action
    Program['order'] = order

